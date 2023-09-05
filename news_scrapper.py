import subprocess

# List of required packages
required_packages = ["requests", "beautifulsoup4", "openpyxl", "lxml"]

# Check if each required package is installed and install it if not
for package in required_packages:
    try:
        import importlib
        importlib.import_module(package)
    except ImportError:
        print(f"Developed By TechView Team .Installing {package}...")
        subprocess.check_call(["pip", "install", package])


import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Function to scrape news articles from a given RSS feed URL
def scrape_news_from_rss(rss_url):
    try:
        print(f"Scraping from URL: {rss_url}")
        response = requests.get(rss_url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'xml')  # Use 'xml' parser for XML
        items = soup.find_all('item')

        news_data = []
        
        # Inside the loop that scrapes data from the RSS feed
        for item in items:
            title = item.title.text.strip() if item.title else "N/A"
            description = item.description.text.strip() if item.description else "N/A"
            link = item.link.text.strip() if item.link else "N/A"
            pubDate = item.pubDate.text.strip() if item.pubDate else "N/A"

            # Handle CDATA sections for title and description
            if item.title and item.title.string:
                title = item.title.string.strip()
            if item.description and item.description.string:
                description = item.description.string.strip()

            # Handle the 'guid' element with 'https://www.hindustantimes.com' prefix
            guid = item.guid.text.strip() if item.guid else "N/A"
            if guid.startswith("https://www.hindustantimes.com"):
                link = guid  # Use 'guid' as the link if it starts with the specified prefix

            # Append the scraped data to the list with the correct keys
            news_data.append({
                'title': title,
                'date': pubDate,
                'description': description,
                'link': link
            })

        return news_data
    except Exception as e:
        print(f"Error scraping from {rss_url}: {str(e)}")
        return []

# Create a directory to store daily news files
output_dir = "daily_news"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Read RSS feed URLs from a text file
with open('rss_urls.txt', 'r') as file:
    rss_urls = file.read().splitlines()

print("Found the following URLs:")
for rss_url in rss_urls:
    print(rss_url)

# Get today's date
today_date = datetime.now().strftime("%Y-%m-%d")

# Create an Excel workbook for today's news or load an existing one
excel_filename = os.path.join(output_dir, f"{today_date}_news.xlsx")
if os.path.exists(excel_filename):
    workbook = load_workbook(excel_filename)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Headline", "Date", "Description", "Link"])  # Add header row for new file

# Create a set to keep track of saved news articles
saved_articles = set()

# Populate the set with existing headlines and dates in the Excel file
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    headline, date, description, link = row
    saved_articles.add((headline, date))

# Loop through the RSS feed URLs and scrape news data
for rss_url in rss_urls:
    news_data = scrape_news_from_rss(rss_url)

    # Write data to the Excel sheet if it's not already saved
    for news_item in news_data:
        title, date, description, link = news_item['title'], news_item['date'], news_item['description'], news_item['link']

        if (title, date) not in saved_articles:
            sheet.append([title, date, description, link])
            saved_articles.add((title, date))

# Save the Excel workbook
workbook.save(excel_filename)

print(f"News for {today_date} saved to {excel_filename}")
